#!/usr/bin/env python3
"""
Build an Excel planning for a given academy from the three DOM dumps
produced by extract_cfjjb.py:

  output/participants_by_team.json  — who is registered, per team
  output/planning.json              — each fight slot on each mat
  output/brackets_by_category.json  — pages per category with fighters & seeds

Join logic:
  athlete (from participants)
    -> bracket category
       -> the page (1..N) on which the athlete sits
          -> planning slot with matching (category, page, mat)
             -> date, time, mat

Output columns:
  Participant | Équipe | Catégorie | Date | Heure | Tapis | Tableaux | Note
  (Tableaux is "N/M" — the athlete's bracket page out of the total pages.)

One row per athlete (no more "same category × N slots" blow-up).
Output lands in output/planning_<academy>.<xlsx|pdf>.

Multiple --input-dir values are allowed: the join runs independently per
directory (each one is a separate CFJJB competition, with its own
categories / pages / mats), then the resulting rows are concatenated and
sorted chronologically. Useful for events split across several
competitions (Gi / No-Gi / Kids Gi / Kids No-Gi).

Usage:
    pip install openpyxl reportlab
    python3 build_planning_xlsx.py --academy "Example Academy"
    python3 build_planning_xlsx.py --academy "Example Academy" --format pdf
    python3 build_planning_xlsx.py --academy "Example Academy" \
        --input-dir output/941 output/942 output/943 output/944 \
        --format pdf --name orleans_2026
"""

from __future__ import annotations

import argparse
import json
import sys
import unicodedata
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------- helpers ----------

def norm(s: str | None) -> str:
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.lower().split())


def load_json(path: Path) -> object:
    if not path.exists():
        print(f"[!] missing: {path}", file=sys.stderr)
        return None
    return json.loads(path.read_text(encoding="utf-8"))


# ---------- shaping ----------

def collect_academy_athletes(participants: list[dict], academy: str) -> list[dict]:
    target = norm(academy)
    out: list[dict] = []
    for team_block in participants or []:
        team = team_block.get("team") or ""
        if target not in norm(team):
            continue
        for ath in team_block.get("athletes", []):
            out.append({
                "name": ath.get("name", ""),
                "team": team,
                "category": ath.get("category", ""),
            })
    return out


def index_brackets(brackets: list[dict]) -> dict[str, dict]:
    return {norm(b.get("raw") or b.get("category", "")): b for b in brackets or []}


def index_planning(planning: list[dict]) -> dict[tuple, list[dict]]:
    """key = (category_norm, page_or_None, mat_norm) -> slots.
       Also stores a fallback key (category_norm, None, None) for single-page cats."""
    idx: dict[tuple, list[dict]] = {}
    for f in planning or []:
        cat = norm(f.get("category", ""))
        if not cat:
            continue
        page = f.get("fight_index")  # scraper field = page number of the tableau
        mat  = norm(f.get("mat", ""))
        idx.setdefault((cat, page, mat), []).append(f)
        idx.setdefault((cat, None, None), []).append(f)  # broad fallback
    return idx


def locate_athlete_in_bracket(name: str, bracket: dict | None) -> tuple[int | None, int | None, str, dict | None]:
    """Return (seed, page, mat_from_bracket, bracket_entry_for_name)."""
    if not bracket:
        return (None, None, "", None)
    target = norm(name)
    for f in bracket.get("fighters", []) or []:
        if norm(f.get("name")) == target:
            return (f.get("seed"), f.get("page"), f.get("mat") or "", f)
    return (None, None, "", None)


# ---------- writer ----------

HEADERS_BASE = [
    "Participant", "Équipe", "Catégorie",
    "Date", "Heure", "Tapis", "Tableaux", "Note",
]


def make_headers(include_competition: bool) -> list[str]:
    if include_competition:
        return HEADERS_BASE[:-1] + ["Compétition", HEADERS_BASE[-1]]
    return list(HEADERS_BASE)

BELT_FILLS = {
    "blanche":  "FFFFFF",
    "bleue":    "DBEAFE",
    "violette": "E9D5FF",
    "marron":   "D7B899",
    "noire":    "374151",
}


def format_tableaux(r: dict) -> str:
    page = r.get("page")
    total = r.get("page_total")
    if page is not None and total:
        return f"{page}/{total}"
    if page is not None:
        return f"{page}"
    if r.get("has_slot"):
        # The planning slot had no "N/M" pill — the bracket is a single
        # page (e.g. P.Seule or a small category). Render as 1/1.
        return "1/1"
    return ""


def row_cells(r: dict, include_competition: bool = False) -> list[str]:
    base = [
        r["name"], r["team"], r["category"],
        r["date"], r["time"], r["mat"],
        format_tableaux(r),
    ]
    if include_competition:
        base.append(r.get("competition") or "")
    base.append(r["note"] or "")
    return base


def write_xlsx(rows: list[dict], out_path: Path, academy: str,
               include_competition: bool = False) -> None:
    headers = make_headers(include_competition)
    wb = Workbook()
    ws = wb.active
    ws.title = "Planning"

    ws.append([f"Planning — {academy}"] + [""] * (len(headers) - 1))
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=2, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        ws.append(row_cells(r, include_competition))
        belt = norm(r.get("belt"))
        fill_hex = BELT_FILLS.get(belt)
        if fill_hex:
            fill = PatternFill("solid", fgColor=fill_hex)
            font_color = "FFFFFF" if belt == "noire" else "000000"
            for col in (1, 2, 3):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.fill = fill
                cell.font = Font(color=font_color)

    widths_base = [28, 24, 38, 12, 8, 12, 10, 28]
    if include_competition:
        widths = widths_base[:-1] + [18, widths_base[-1]]
    else:
        widths = widths_base
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{ws.max_row}"
    wb.save(out_path)


def write_pdf(rows: list[dict], out_path: Path, academy: str,
              include_competition: bool = False) -> None:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )

    def hex_to_color(hex_str: str) -> colors.Color:
        return colors.HexColor("#" + hex_str)

    headers = make_headers(include_competition)

    page_size = landscape(A4)
    doc = SimpleDocTemplate(
        str(out_path),
        pagesize=page_size,
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=10 * mm, bottomMargin=10 * mm,
        title=f"Planning — {academy}",
    )

    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    title_style.fontSize = 14
    cell_style = styles["BodyText"]
    cell_style.fontSize = 8
    cell_style.leading = 10

    def P(text: str) -> Paragraph:
        # Wrap long cells (category, note) instead of overflowing.
        safe = (text or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        return Paragraph(safe, cell_style)

    note_col = len(headers) - 1  # last column is always "Note"
    data: list[list] = [headers]
    for r in rows:
        cells = row_cells(r, include_competition)
        row = []
        for i, val in enumerate(cells):
            # Wrap text in the first three columns and the last (Note)
            # + the competition column if present — these can be long.
            if i in (0, 1, 2) or i == note_col or (include_competition and i == note_col - 1):
                row.append(P(val))
            else:
                row.append(val)
        data.append(row)

    # Column widths (mm) — sum ~= 277mm (A4 landscape minus margins).
    if include_competition:
        # Squeeze the note column to fit Compétition in.
        col_widths = [46, 36, 58, 20, 14, 22, 16, 22, 43]
    else:
        col_widths = [48, 40, 62, 20, 14, 22, 16, 55]
    col_widths = [w * mm for w in col_widths]

    table = Table(data, colWidths=col_widths, repeatRows=1)

    # Centered columns: Date/Heure/Tapis/Tableaux are fixed-width fields.
    # Their column indices are 3..6 in both layouts.
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), hex_to_color("1E3A8A")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 9),
        ("ALIGN",      (0, 0), (-1, 0), "CENTER"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",       (0, 0), (-1, -1), 0.25, colors.grey),
        ("FONTSIZE",   (0, 1), (-1, -1), 8),
        ("ALIGN",      (3, 1), (6, -1), "CENTER"),
    ])

    for i, r in enumerate(rows, start=1):
        belt = norm(r.get("belt"))
        fill_hex = BELT_FILLS.get(belt)
        if not fill_hex:
            continue
        fill = hex_to_color(fill_hex)
        font_color = colors.white if belt == "noire" else colors.black
        style.add("BACKGROUND", (0, i), (2, i), fill)
        style.add("TEXTCOLOR",  (0, i), (2, i), font_color)

    table.setStyle(style)

    doc.build([
        Paragraph(f"Planning — {academy}", title_style),
        Spacer(1, 4 * mm),
        table,
    ])


# ---------- main ----------

def build_rows_for_dir(
    in_dir: Path,
    academy: str,
    competition_label: str,
    debug: bool = False,
    debug_target: str = "",
) -> tuple[list[dict], int]:
    """Run the full join (participants × brackets × planning) inside one
    competition directory. Returns (rows, unresolved_count). Each row gets
    the competition_label stamped on it so merged outputs stay traceable."""
    participants = load_json(in_dir / "participants_by_team.json") or []
    planning     = load_json(in_dir / "planning.json") or []
    brackets     = load_json(in_dir / "brackets_by_category.json") or []

    athletes = collect_academy_athletes(participants, academy)

    bracket_idx  = index_brackets(brackets)
    planning_idx = index_planning(planning)

    print(
        f"[+] [{competition_label}] {len(athletes)} athlete(s) for '{academy}' | "
        f"{len(bracket_idx)} bracket category(ies) | "
        f"{sum(len(v) for k, v in planning_idx.items() if k[1] is not None)} located planning slot(s).",
        file=sys.stderr,
    )

    def trace(ath: dict, *parts: str) -> None:
        if not debug:
            return
        if debug_target and debug_target not in norm(ath.get("name", "")):
            return
        print("  " + " | ".join(parts), file=sys.stderr)

    rows: list[dict] = []
    unresolved = 0
    for ath in athletes:
        cat_key = norm(ath["category"])
        bracket = bracket_idx.get(cat_key)
        weight_limit = (bracket or {}).get("weight_limit") or ""
        belt = (bracket or {}).get("belt") or ath["category"].split(" - ", 1)[0]
        page_total = (bracket or {}).get("page_total")

        seed, page, bracket_mat, _ = locate_athlete_in_bracket(ath["name"], bracket)

        if debug and (not debug_target or debug_target in norm(ath.get("name", ""))):
            print(f"[{competition_label}] [{ath['name']}] team={ath['team']!r} "
                  f"cat={ath['category']!r} cat_key={cat_key!r} "
                  f"bracket_found={bracket is not None} "
                  f"page={page} mat={bracket_mat!r} seed={seed}",
                  file=sys.stderr)

        # Priority 1: exact match (category, page, mat).
        slot = None
        priority = None
        if page is not None and bracket_mat:
            key1 = (cat_key, page, norm(bracket_mat))
            p1_hits = planning_idx.get(key1, [])
            trace(ath, f"P1 key={key1}", f"hits={len(p1_hits)}")
            for s in p1_hits:
                slot = s; priority = 1; break
        # Priority 2: (category, page, any mat) — rare case of mat typo.
        if slot is None and page is not None:
            for (cat, pg, mat_k), slots in planning_idx.items():
                if cat == cat_key and pg == page and slots:
                    trace(ath, f"P2 any-mat match: (cat={cat_key!r}, page={page}, mat={mat_k!r})")
                    slot = slots[0]; priority = 2; break
        # Priority 3: single-page or single-slot category.
        if slot is None:
            broad = planning_idx.get((cat_key, None, None), [])
            if broad:
                # Prefer the one on the bracket mat, else the earliest.
                broad_sorted = sorted(
                    broad,
                    key=lambda s: (
                        0 if norm(s.get("mat")) == norm(bracket_mat) else 1,
                        s.get("date") or "", s.get("time") or "",
                    ),
                )
                slot = broad_sorted[0]
                priority = 3
                trace(ath, f"P3 broad fallback: {len(broad)} slot(s) for category, "
                           f"picked {slot.get('date')} {slot.get('time')} {slot.get('mat')} "
                           f"fight={slot.get('fight_index')}/{slot.get('fight_total')}")

        trace(ath, f"=> priority={priority} slot="
                   + (f"{slot.get('date')} {slot.get('time')} {slot.get('mat')} "
                      f"fight={slot.get('fight_index')}/{slot.get('fight_total')}"
                      if slot else "None"))

        if slot is None:
            unresolved += 1
            # Build a helpful note so the user can see *why* it's missing.
            hints = []
            if bracket is None:
                hints.append("catégorie absente des tableaux")
            elif page is None:
                hints.append("athlète non localisé dans le tableau")
            if not planning_idx.get((cat_key, None, None)):
                hints.append("catégorie absente du planning")
            rows.append({
                "name": ath["name"], "team": ath["team"],
                "category": ath["category"], "weight_limit": weight_limit,
                "seed": seed, "belt": belt,
                "date": "", "time": "", "mat": bracket_mat,
                "page": page, "page_total": page_total,
                "note": "; ".join(hints) or "pas de créneau trouvé",
                "competition": competition_label,
            })
            continue

        rows.append({
            "name": ath["name"], "team": ath["team"],
            "category": ath["category"], "weight_limit": weight_limit,
            "seed": seed, "belt": belt,
            "date": slot.get("date") or "",
            "time": slot.get("time") or "",
            "mat":  slot.get("mat")  or bracket_mat,
            "page": slot.get("fight_index") if slot.get("fight_index") is not None else page,
            "page_total": slot.get("fight_total") if slot.get("fight_total") is not None else page_total,
            "note": slot.get("note") or "",
            "has_slot": True,
            "competition": competition_label,
        })

    return rows, unresolved


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--academy", required=True, help="Academy to build the planning for (e.g. 'Example Academy').")
    ap.add_argument("--input-dir", nargs="+", default=["output"],
                    help="One or more folders with the three JSON files. "
                         "Pass multiple to merge several competitions of the "
                         "same event (e.g. --input-dir output/941 output/942).")
    ap.add_argument("--format", choices=("xlsx", "pdf"), default="xlsx",
                    help="Output format (default: xlsx).")
    ap.add_argument("--name", default=None,
                    help="Output filename stem placed under the first --input-dir "
                         "(e.g. --name mon_planning -> <input-dir>/mon_planning.<format>). "
                         "Any extension you include is stripped.")
    ap.add_argument("--out", default=None,
                    help="Full output path, overrides --name and --input-dir "
                         "(default: <first-input-dir>/planning_<academy>.<format>).")
    ap.add_argument("--debug", action="store_true",
                    help="Print per-athlete lookup trace (bracket found, page/mat, which priority matched).")
    ap.add_argument("--only", default=None,
                    help="When used with --debug, restrict the trace to athletes whose name contains this substring.")
    args = ap.parse_args()

    in_dirs = [Path(d) for d in args.input_dir]
    debug_target = norm(args.only) if args.only else ""

    all_rows: list[dict] = []
    total_unresolved = 0
    for d in in_dirs:
        label = d.name or str(d)
        rows, unresolved = build_rows_for_dir(
            d, args.academy, label,
            debug=args.debug, debug_target=debug_target,
        )
        all_rows.extend(rows)
        total_unresolved += unresolved

    if not all_rows:
        print(f"[!] no athletes matched academy ~ '{args.academy}' across "
              f"{len(in_dirs)} input dir(s).", file=sys.stderr)
        return 1

    all_rows.sort(key=lambda r: (str(r["date"]), str(r["time"]), str(r["mat"]), str(r["name"])))

    include_competition = len(in_dirs) > 1

    first_dir = in_dirs[0]
    if args.out:
        out_path = Path(args.out)
    elif args.name:
        stem = Path(args.name).stem or args.name
        out_path = first_dir / f"{stem}.{args.format}"
    else:
        slug = "".join(c if c.isalnum() else "_" for c in args.academy).strip("_")
        out_path = first_dir / f"planning_{slug}.{args.format}"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    if args.format == "pdf":
        write_pdf(all_rows, out_path, args.academy, include_competition=include_competition)
    else:
        write_xlsx(all_rows, out_path, args.academy, include_competition=include_competition)
    print(f"[+] wrote {len(all_rows)} row(s) ({total_unresolved} unresolved) -> {out_path}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
